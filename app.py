import io
import os
import re
import hashlib
from typing import Dict, List

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# =========================
# Parsing helpers
# =========================

RE_NON_NUM = re.compile(r"[^0-9\-,.]+")

def parse_number(x):
    """Parse numeric strings like '1.234,56' or '278680.15998' -> float."""
    if pd.isna(x):
        return np.nan
    if isinstance(x, (int, float, np.integer, np.floating)):
        return float(x)

    s = str(x).strip()
    if s == "":
        return np.nan

    s = RE_NON_NUM.sub("", s)

    # Indonesian format: 1.234,56 -> 1234.56
    if s.count(",") == 1 and s.count(".") >= 1:
        s = s.replace(".", "").replace(",", ".")
    # comma decimal: 12,34 -> 12.34
    elif s.count(",") == 1 and s.count(".") == 0:
        s = s.replace(",", ".")

    try:
        return float(s)
    except Exception:
        return np.nan


def trunc_int(x):
    """Truncate decimals (hapus belakang koma), e.g. 1234.99 -> 1234."""
    if pd.isna(x):
        return np.nan
    try:
        return int(np.trunc(float(x)))
    except Exception:
        return np.nan


def series_trunc_int(s: pd.Series) -> pd.Series:
    """Vectorized truncate -> nullable Int64."""
    return s.apply(trunc_int).astype("Int64")


def format_dot_thousands(x) -> str:
    """Format number to '278.680' (dot thousands), no decimals."""
    if pd.isna(x):
        return ""
    try:
        v = int(x)
        return f"{v:,}".replace(",", ".")
    except Exception:
        return ""


def id_to_str(x) -> str:
    """Convert IDs that might come as float/str into clean digit string."""
    if pd.isna(x):
        return ""
    s = str(x).strip()
    s = re.sub(r"\.0$", "", s)
    s = re.sub(r"\D+", "", s)
    return s


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]

    # common typos
    rename_map = {
        "Nama Barange": "Nama Barang",
        "Kampanye Partnerr": "Kampanye Partner",
        "Status Pemebelian": "Status Pembelian",
    }
    for k, v in rename_map.items():
        if k in df.columns and v not in df.columns:
            df = df.rename(columns={k: v})
    return df


def ensure_id_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Robust: cari kolom shop id & item id dari nama kolom (case-insensitive + pattern).
    Output: memastikan ada 'ID Toko' dan/atau 'ID Barang' jika ditemukan.
    """
    df = df.copy()
    cols = list(df.columns)
    lower = {c: c.lower() for c in cols}

    def pick_col(patterns: List[str]) -> str:
        for c in cols:
            cl = lower[c]
            for p in patterns:
                if re.search(p, cl):
                    return c
        return ""

    # kandidat shop id
    if "ID Toko" not in df.columns:
        shop_col = pick_col([
            r"\bid\s*toko\b",
            r"\bshop\s*id\b",
            r"\bid\s*shop\b",
            r"\bshopid\b",
            r"\bidshop\b",
            r"\bseller\s*id\b",
            r"\bshop_id\b",
        ])
        if shop_col:
            df = df.rename(columns={shop_col: "ID Toko"})

    # kandidat item id
    if "ID Barang" not in df.columns:
        item_col = pick_col([
            r"\bid\s*barang\b",
            r"\bitem\s*id\b",
            r"\bid\s*item\b",
            r"\bitemid\b",
            r"\biditem\b",
            r"\bproduct\s*id\b",
            r"\bid\s*produk\b",
            r"\bitem_id\b",
        ])
        if item_col:
            df = df.rename(columns={item_col: "ID Barang"})

    return df


def add_derived_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    # date columns
    for col in ["Waktu Pemesanan", "Waktu Terselesaikan", "Waktu Klik"]:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")

    if "Waktu Pemesanan" in df.columns:
        df["Tanggal"] = df["Waktu Pemesanan"].dt.date

    # numeric columns parse
    num_cols = [
        "Harga(Rp)",
        "Jumlah",
        "Nilai Pembelian(Rp)",
        "Jumlah Pengembalian Dana(Rp)",
        "Komisi Bersih Affiliate (Rp)",
        "Total Komisi per Pesanan(Rp)",
        "Total Komisi per Produk(Rp)",
        "Komisi Shopee per Pesanan(Rp)",
        "Komisi XTRA per Pesanan(Rp)",
    ]
    for c in num_cols:
        if c in df.columns:
            df[c] = df[c].apply(parse_number)

    # truncate decimals for money + qty
    money_like = [
        "Harga(Rp)",
        "Nilai Pembelian(Rp)",
        "Jumlah Pengembalian Dana(Rp)",
        "Komisi Bersih Affiliate (Rp)",
        "Total Komisi per Pesanan(Rp)",
        "Total Komisi per Produk(Rp)",
        "Komisi Shopee per Pesanan(Rp)",
        "Komisi XTRA per Pesanan(Rp)",
    ]
    for c in money_like:
        if c in df.columns:
            df[c] = series_trunc_int(df[c])

    if "Jumlah" in df.columns:
        df["Jumlah"] = series_trunc_int(df["Jumlah"].fillna(0))

    # status flags
    if "Status Pesanan" in df.columns:
        df["Status Pesanan"] = df["Status Pesanan"].astype(str).str.strip()
        df["is_pending"] = df["Status Pesanan"].str.lower().eq("tertunda")
        df["is_completed"] = df["Status Pesanan"].str.lower().isin(["selesai", "dibayarkan", "completed"])
    else:
        df["is_pending"] = False
        df["is_completed"] = False

    # link produk
    if "ID Toko" in df.columns and "ID Barang" in df.columns:
        df["ID Toko"] = df["ID Toko"].apply(id_to_str)
        df["ID Barang"] = df["ID Barang"].apply(id_to_str)

        shop = df["ID Toko"]
        item = df["ID Barang"]
        df["Link Produk"] = np.where(
            (shop != "") & (item != ""),
            "https://shopee.co.id/product/" + shop + "/" + item,
            ""
        )
    else:
        df["Link Produk"] = ""

    return df


def read_csv_bytes(raw: bytes) -> pd.DataFrame:
    # dtype=str supaya ID tidak jadi float / scientific
    for enc in ["utf-8-sig", "utf-8", "latin1"]:
        try:
            return pd.read_csv(io.BytesIO(raw), encoding=enc, dtype=str)
        except Exception:
            continue
    return pd.read_csv(io.BytesIO(raw), encoding_errors="ignore", dtype=str)


def file_md5(raw: bytes) -> str:
    return hashlib.md5(raw).hexdigest()


@st.cache_data(show_spinner=False)
def parse_one_file(raw: bytes, filename: str) -> pd.DataFrame:
    df = read_csv_bytes(raw)
    df = normalize_columns(df)
    df = ensure_id_columns(df)          # << penting: deteksi ID toko/barang
    df = add_derived_columns(df)
    akun = os.path.splitext(os.path.basename(filename))[0]
    df["Akun"] = akun
    return df


@st.cache_data(show_spinner=False)
def parse_many(files_payload: List[dict]) -> pd.DataFrame:
    frames = []
    for item in files_payload:
        frames.append(parse_one_file(item["raw"], item["name"]))
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


# =========================
# Demand metric logic
# =========================

def demand_metric_label(demand_mode: str) -> str:
    return "Orders (unik)" if demand_mode == "Pesanan (Orders unik)" else "Produk terjual (Items/Jumlah)"


def compute_demand_agg(df: pd.DataFrame, group_cols: List[str], demand_mode: str) -> pd.DataFrame:
    order_id_col = "ID Pemesanan" if "ID Pemesanan" in df.columns else None

    gb = df.groupby(group_cols, dropna=False)

    orders = gb[order_id_col].nunique() if order_id_col else gb.size()
    items = gb["Jumlah"].sum() if "Jumlah" in df.columns else gb.size()

    res = pd.DataFrame({
        "Orders (unik)": orders,
        "Items (Jumlah)": items,
    }).reset_index()

    # context
    if "Nilai Pembelian(Rp)" in df.columns:
        res["GMV (Rp)"] = gb["Nilai Pembelian(Rp)"].sum().values
    if "Komisi Bersih Affiliate (Rp)" in df.columns:
        res["Komisi Bersih (Rp)"] = gb["Komisi Bersih Affiliate (Rp)"].sum().values

    res["Demand"] = res["Orders (unik)"] if demand_mode == "Pesanan (Orders unik)" else res["Items (Jumlah)"]

    # pastikan integer (truncate) di agregasi
    for c in ["Orders (unik)", "Items (Jumlah)", "Demand", "GMV (Rp)", "Komisi Bersih (Rp)"]:
        if c in res.columns:
            res[c] = series_trunc_int(pd.to_numeric(res[c], errors="coerce"))

    return res


# =========================
# Analytics tables
# =========================

def overview_table(df: pd.DataFrame) -> pd.DataFrame:
    date_min = df["Tanggal"].min() if "Tanggal" in df.columns else None
    date_max = df["Tanggal"].max() if "Tanggal" in df.columns else None

    order_id_col = "ID Pemesanan" if "ID Pemesanan" in df.columns else None
    unique_orders = df[order_id_col].nunique(dropna=True) if order_id_col else len(df)
    items = int(df["Jumlah"].sum()) if "Jumlah" in df.columns else len(df)

    total_gmv = int(df["Nilai Pembelian(Rp)"].sum()) if "Nilai Pembelian(Rp)" in df.columns else 0
    total_comm = int(df["Komisi Bersih Affiliate (Rp)"].sum()) if "Komisi Bersih Affiliate (Rp)" in df.columns else 0
    pending_cnt = int(df["is_pending"].sum()) if "is_pending" in df.columns else 0

    out = [
        ("Periode", f"{date_min} s/d {date_max}"),
        ("Jumlah baris", int(len(df))),
        ("Akun unik", int(df["Akun"].nunique()) if "Akun" in df.columns else 1),
        ("Pesanan unik (Orders)", int(unique_orders)),
        ("Produk terjual (Items/Jumlah)", items),
        ("Total Nilai Pembelian (GMV)", total_gmv),
        ("Total Komisi Bersih Affiliate", total_comm),
        ("Jumlah baris status Tertunda", pending_cnt),
    ]
    return pd.DataFrame(out, columns=["Metric", "Value"])


def daily_totals(df: pd.DataFrame) -> pd.DataFrame:
    if "Tanggal" not in df.columns:
        return pd.DataFrame()

    res = compute_demand_agg(df, ["Tanggal"], demand_mode="Pesanan (Orders unik)")

    if "GMV (Rp)" in res.columns:
        aov = np.where(res["Orders (unik)"].astype(float) > 0,
                       res["GMV (Rp)"].astype(float) / res["Orders (unik)"].astype(float),
                       np.nan)
        res["AOV (Rp)"] = series_trunc_int(pd.Series(aov))

    if "Komisi Bersih (Rp)" in res.columns:
        kpo = np.where(res["Orders (unik)"].astype(float) > 0,
                       res["Komisi Bersih (Rp)"].astype(float) / res["Orders (unik)"].astype(float),
                       np.nan)
        res["Komisi / Order (Rp)"] = series_trunc_int(pd.Series(kpo))

    return res.sort_values("Tanggal")


def summary_by_category(df: pd.DataFrame, level_col: str, demand_mode: str) -> pd.DataFrame:
    if level_col not in df.columns:
        return pd.DataFrame()

    res = compute_demand_agg(df, [level_col], demand_mode=demand_mode).rename(columns={level_col: "Kategori"})
    total = float(res["Demand"].sum()) if res["Demand"].notna().any() else 0.0
    res["Share Demand"] = np.where(total > 0, res["Demand"].astype(float) / total, np.nan)
    return res.sort_values("Demand", ascending=False)


def top_products(df: pd.DataFrame, demand_mode: str, top_n: int = 30) -> pd.DataFrame:
    keys = []
    if "ID Toko" in df.columns:
        keys.append("ID Toko")
    if "ID Barang" in df.columns:
        keys.append("ID Barang")
    if "Nama Barang" in df.columns:
        keys.append("Nama Barang")

    if ("ID Barang" not in df.columns):
        return pd.DataFrame()

    res = compute_demand_agg(df, keys, demand_mode=demand_mode)

    # merge link jika ada
    if "ID Toko" in df.columns and "ID Barang" in df.columns and "Link Produk" in df.columns:
        link_map = (
            df.loc[df["Link Produk"] != "", ["ID Toko", "ID Barang", "Link Produk"]]
              .drop_duplicates(subset=["ID Toko", "ID Barang"])
        )
        res = res.merge(link_map, on=["ID Toko", "ID Barang"], how="left")

    return res.sort_values("Demand", ascending=False).head(top_n)


def top_stores(df: pd.DataFrame, demand_mode: str, top_n: int = 30) -> pd.DataFrame:
    if "Nama Toko" not in df.columns:
        return pd.DataFrame()
    res = compute_demand_agg(df, ["Nama Toko"], demand_mode=demand_mode)
    return res.sort_values("Demand", ascending=False).head(top_n)


def winning_l1_daily(df: pd.DataFrame, l1_col: str, demand_mode: str) -> pd.DataFrame:
    if "Tanggal" not in df.columns or l1_col not in df.columns:
        return pd.DataFrame()

    g = compute_demand_agg(df, ["Tanggal", l1_col], demand_mode=demand_mode)
    g = g.sort_values(["Tanggal", "Demand"], ascending=[True, False])

    winner = g.groupby("Tanggal").head(1).copy().rename(columns={l1_col: "Winning L1"})
    total_day = g.groupby("Tanggal")["Demand"].sum().reset_index().rename(columns={"Demand": "Total Demand Hari Itu"})
    winner = winner.merge(total_day, on="Tanggal", how="left")
    winner["Share"] = np.where(
        winner["Total Demand Hari Itu"].astype(float) > 0,
        winner["Demand"].astype(float) / winner["Total Demand Hari Itu"].astype(float),
        np.nan
    )
    return winner.sort_values("Tanggal")


def daily_top_k_category(df: pd.DataFrame, group_col: str, demand_mode: str, top_k: int) -> pd.DataFrame:
    if "Tanggal" not in df.columns or group_col not in df.columns:
        return pd.DataFrame()

    g = compute_demand_agg(df, ["Tanggal", group_col], demand_mode=demand_mode).rename(columns={group_col: "Item"})
    g["Rank"] = g.groupby("Tanggal")["Demand"].rank(method="first", ascending=False)
    return g[g["Rank"] <= top_k].sort_values(["Tanggal", "Rank"])


def daily_top_k_products(df: pd.DataFrame, demand_mode: str, top_k: int) -> pd.DataFrame:
    if "Tanggal" not in df.columns or "ID Barang" not in df.columns:
        return pd.DataFrame()

    keys = ["Tanggal"]
    if "ID Toko" in df.columns:
        keys += ["ID Toko"]
    keys += ["ID Barang"]
    if "Nama Barang" in df.columns:
        keys += ["Nama Barang"]

    g = compute_demand_agg(df, keys, demand_mode=demand_mode)

    if "ID Toko" in df.columns and "Link Produk" in df.columns:
        link_map = (
            df.loc[df["Link Produk"] != "", ["ID Toko", "ID Barang", "Link Produk"]]
              .drop_duplicates(subset=["ID Toko", "ID Barang"])
        )
        g = g.merge(link_map, on=["ID Toko", "ID Barang"], how="left")

    g["Rank"] = g.groupby("Tanggal")["Demand"].rank(method="first", ascending=False)
    return g[g["Rank"] <= top_k].sort_values(["Tanggal", "Rank"])


# =========================
# Excel export (no decimals)
# =========================

def autosize_columns(ws, max_width: int = 60):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            try:
                v = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(v))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_width, max(10, max_len + 2))


def style_header(ws, header_row: int = 1):
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[header_row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = align
    ws.freeze_panes = ws["A2"]


def apply_number_formats(ws, int_cols=None, percent_cols=None, date_cols=None):
    int_cols = int_cols or []
    percent_cols = percent_cols or []
    date_cols = date_cols or []

    header = [c.value for c in ws[1]]
    col_index = {name: i + 1 for i, name in enumerate(header) if name is not None}

    for name in int_cols:
        if name in col_index:
            idx = col_index[name]
            for r in range(2, ws.max_row + 1):
                ws.cell(r, idx).number_format = "#,##0"

    for name in percent_cols:
        if name in col_index:
            idx = col_index[name]
            for r in range(2, ws.max_row + 1):
                ws.cell(r, idx).number_format = "0.0%"

    for name in date_cols:
        if name in col_index:
            idx = col_index[name]
            for r in range(2, ws.max_row + 1):
                ws.cell(r, idx).number_format = "yyyy-mm-dd"


def export_excel_bytes(tables: Dict[str, pd.DataFrame]) -> bytes:
    # pastikan integer columns benar-benar int sebelum export
    int_like_cols = [
        "Orders (unik)", "Items (Jumlah)", "Demand", "Total Demand Hari Itu",
        "GMV (Rp)", "Komisi Bersih (Rp)", "AOV (Rp)", "Komisi / Order (Rp)", "Rank"
    ]

    export_tables = {}
    for name, df in tables.items():
        d = df.copy()
        for c in int_like_cols:
            if c in d.columns:
                vals = pd.to_numeric(d[c], errors="coerce").fillna(0).astype(float)
                d[c] = np.trunc(vals).astype("int64")
        export_tables[name] = d

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for sheet, table in export_tables.items():
            table.to_excel(writer, sheet_name=sheet[:31], index=False)

    buf.seek(0)
    wb = load_workbook(buf)

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        style_header(ws, 1)
        autosize_columns(ws)
        apply_number_formats(
            ws,
            int_cols=int_like_cols,
            percent_cols=["Share Demand", "Share"],
            date_cols=["Tanggal"],
        )

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# =========================
# Streamlit UI
# =========================

st.set_page_config(page_title="Affiliate Demand Winning Analyzer", layout="wide")
st.title("📊 Affiliate Winning Analyzer (Winning = Demand Konsumen)")

with st.sidebar:
    st.header("Upload")
    uploaded = st.file_uploader("Upload CSV (bisa banyak akun)", type=["csv"], accept_multiple_files=True)

    st.divider()
    st.header("Winning Metric (Demand)")
    demand_mode = st.selectbox(
        "Winning berdasarkan:",
        ["Pesanan (Orders unik)", "Produk terjual (Items/Jumlah)"],
        index=0
    )
    demand_col_name = demand_metric_label(demand_mode)

    st.divider()
    st.header("Filter")
    only_pending = st.checkbox("Hanya Status Pesanan = Tertunda", value=False)
    only_completed = st.checkbox("Hanya pesanan selesai/dibayarkan (flag)", value=False)

    st.divider()
    st.header("Top-N")
    top_products_n = st.slider("Top Products (overall)", 5, 200, 30, 5)
    top_stores_n = st.slider("Top Stores (overall)", 5, 200, 30, 5)

if not uploaded:
    st.info("Upload satu atau beberapa CSV dulu ya.")
    st.stop()

files_payload = [{"name": f.name, "raw": f.getvalue(), "md5": hashlib.md5(f.getvalue()).hexdigest()} for f in uploaded]
df = parse_many(files_payload)

if df.empty:
    st.error("Data kosong / gagal dibaca.")
    st.stop()

work = df.copy()
if only_pending and "is_pending" in work.columns:
    work = work[work["is_pending"] == True]
if only_completed and "is_completed" in work.columns:
    work = work[work["is_completed"] == True]

# Date filter
if "Tanggal" in work.columns and work["Tanggal"].notna().any():
    dmin = pd.to_datetime(work["Tanggal"].min())
    dmax = pd.to_datetime(work["Tanggal"].max())
    c1, c2, _ = st.columns([1, 1, 2])
    with c1:
        start = st.date_input("Tanggal mulai", value=dmin.date(), min_value=dmin.date(), max_value=dmax.date())
    with c2:
        end = st.date_input("Tanggal akhir", value=dmax.date(), min_value=dmin.date(), max_value=dmax.date())
    work = work[(pd.to_datetime(work["Tanggal"]) >= pd.to_datetime(start)) &
                (pd.to_datetime(work["Tanggal"]) <= pd.to_datetime(end))]

# Tables
tables: Dict[str, pd.DataFrame] = {}
tables["Overview"] = overview_table(work)
tables["Daily Totals"] = daily_totals(work)
tables["Summary L1"] = summary_by_category(work, "L1 Kategori Global", demand_mode=demand_mode)
tables["Summary L2"] = summary_by_category(work, "L2 Kategori Global", demand_mode=demand_mode)
tables["Top Products"] = top_products(work, demand_mode=demand_mode, top_n=top_products_n)
tables["Top Stores"] = top_stores(work, demand_mode=demand_mode, top_n=top_stores_n)
tables["Winning L1 Daily"] = winning_l1_daily(work, "L1 Kategori Global", demand_mode=demand_mode)
tables["Daily Top L1 Top3"] = daily_top_k_category(work, "L1 Kategori Global", demand_mode=demand_mode, top_k=3)
tables["Daily Top Products Top5"] = daily_top_k_products(work, demand_mode=demand_mode, top_k=5)


def display_df_with_dot(df_in: pd.DataFrame, cols: List[str]) -> pd.DataFrame:
    d = df_in.copy()
    for c in cols:
        if c in d.columns:
            d[c] = d[c].apply(format_dot_thousands)
    return d


tab1, tab2, tab3 = st.tabs(["📌 Overview", "📈 Harian", "🏆 Winning (Demand)"])

with tab1:
    st.subheader("Overview")
    st.dataframe(tables["Overview"], use_container_width=True)
    st.caption(f"Rows: {len(work):,} | Akun unik: {work['Akun'].nunique() if 'Akun' in work.columns else 1:,}")

with tab2:
    st.subheader("Daily Totals (tampilan titik, tanpa desimal)")
    dt = tables["Daily Totals"]
    st.dataframe(display_df_with_dot(dt, ["Orders (unik)", "Items (Jumlah)", "GMV (Rp)", "Komisi Bersih (Rp)", "AOV (Rp)", "Komisi / Order (Rp)"]),
                 use_container_width=True)

    if not dt.empty:
        dt_idx = dt.set_index("Tanggal")
        c1, c2 = st.columns(2)
        with c1:
            st.markdown(f"**Trend Demand** ({demand_col_name})")
            st.line_chart(dt_idx[["Orders (unik)"]].astype(float) if demand_mode == "Pesanan (Orders unik)"
                          else dt_idx[["Items (Jumlah)"]].astype(float))
        with c2:
            cols = [c for c in ["GMV (Rp)", "Komisi Bersih (Rp)"] if c in dt_idx.columns]
            if cols:
                st.markdown("**Trend GMV / Komisi (konteks)**")
                st.line_chart(dt_idx[cols].astype(float))

with tab3:
    c1, c2 = st.columns(2)

    with c1:
        st.subheader(f"Summary L1 — Ranking by {demand_col_name}")
        s1 = tables["Summary L1"]
        st.dataframe(display_df_with_dot(s1, ["Orders (unik)", "Items (Jumlah)", "Demand", "GMV (Rp)", "Komisi Bersih (Rp)"]).head(50),
                     use_container_width=True)

    with c2:
        st.subheader("Winning L1 per Hari")
        wl = tables["Winning L1 Daily"]
        st.dataframe(display_df_with_dot(wl, ["Orders (unik)", "Items (Jumlah)", "Demand", "Total Demand Hari Itu", "GMV (Rp)", "Komisi Bersih (Rp)"]),
                     use_container_width=True)

    st.subheader("Top Products / Stores")
    c3, c4 = st.columns(2)
    with c3:
        tp = display_df_with_dot(tables["Top Products"], ["Orders (unik)", "Items (Jumlah)", "Demand", "GMV (Rp)", "Komisi Bersih (Rp)"])
        if "Link Produk" in tp.columns:
            st.dataframe(tp, use_container_width=True,
                         column_config={"Link Produk": st.column_config.LinkColumn("Link Produk")})
        else:
            st.dataframe(tp, use_container_width=True)

    with c4:
        ts = display_df_with_dot(tables["Top Stores"], ["Orders (unik)", "Items (Jumlah)", "Demand", "GMV (Rp)", "Komisi Bersih (Rp)"])
        st.dataframe(ts, use_container_width=True)

    st.subheader("Daily Top")
    c5, c6 = st.columns(2)
    with c5:
        d1 = display_df_with_dot(tables["Daily Top L1 Top3"], ["Orders (unik)", "Items (Jumlah)", "Demand", "GMV (Rp)", "Komisi Bersih (Rp)", "Rank"])
        st.dataframe(d1, use_container_width=True)

    with c6:
        dp = display_df_with_dot(tables["Daily Top Products Top5"], ["Orders (unik)", "Items (Jumlah)", "Demand", "GMV (Rp)", "Komisi Bersih (Rp)", "Rank"])
        if "Link Produk" in dp.columns:
            st.dataframe(dp, use_container_width=True,
                         column_config={"Link Produk": st.column_config.LinkColumn("Link Produk")})
        else:
            st.dataframe(dp, use_container_width=True)

st.divider()
st.subheader("⬇️ Download Excel Report (tanpa desimal)")
excel_bytes = export_excel_bytes(tables)
st.download_button(
    "Download Excel Report",
    data=excel_bytes,
    file_name="affiliate_demand_winning_report.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)
